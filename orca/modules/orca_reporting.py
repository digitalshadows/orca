import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter


def fix_columns(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value


def add_vuln_ws(orca_dbconn, wb):
    row = 1
    col = 1
    results = orca_dbconn.get_all_vuln_entries()

    ws = wb.create_sheet(title="Vulns")

    c = ws.cell(row=row, column=col)
    c.value = 'Vuln ID'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 1)
    c.value = 'Host ID'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 2)
    c.value = 'IP addr'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 3)
    c.value = 'SHODAN hostname'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 4)
    c.value = 'Module'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 5)
    c.value = 'CPE'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 6)
    c.value = 'CVE'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 7)
    c.value = 'CVSS'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 8)
    c.value = 'Verified'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 9)
    c.value = 'Summary'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 10)
    c.value = 'Exploit-DB Reference'
    c.font = Font(bold=True)

    row += 1

    for result in results:
        c = ws.cell(row=row, column=col)
        c.value = str(result['vuln_id'])

        c = ws.cell(row=row, column=col + 1)
        c.value = str(result['host_id'])

        c = ws.cell(row=row, column=col + 2)
        c.value = str(result['ipaddr'])

        c = ws.cell(row=row, column=col + 3)
        if len(result['shodan_hostname']) > 1:
            c.value = ','.join(result['shodan_hostname'])
        elif len(result['shodan_hostname']) == 1:
            c.value = result['shodan_hostname'][0]

        c = ws.cell(row=row, column=col + 4)
        c.value = str(result['module'])

        c = ws.cell(row=row, column=col + 5)
        if len(result['cpe']) > 1:
            c.value = ','.join(result['cpe'])
        else:
            c.value = result['cpe'][0]

        c = ws.cell(row=row, column=col + 6)
        c.value = str(result['cve'])

        c = ws.cell(row=row, column=col + 7)
        c.value = str(result['cvss'])

        c = ws.cell(row=row, column=col + 8)
        c.value = str(result['verified'])

        c = ws.cell(row=row, column=col + 9)
        c.value = str(result['summary'])
        wrap_alignment = Alignment(wrap_text=True)
        c.alignment = wrap_alignment

        if result['exploit']:
            c = ws.cell(row=row, column=col + 10)
            c.value = str(result['exploit_ref'])
        else:
            c = ws.cell(row=row, column=col + 10)
            c.value = "None"
        row += 1

    fix_columns(ws)


def add_dns_ws(orca_dbconn, wb):
    row = 1
    col = 1
    results = orca_dbconn.get_all_dns_entries()

    ws = wb.create_sheet(title="DNS")

    c = ws.cell(row=row, column=col)
    c.value = 'Domain'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 1)
    c.value = 'A'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 2)
    c.value = 'CNAME'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 3)
    c.value = 'TXT'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 4)
    c.value = 'AAAA'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 5)
    c.value = 'MX'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 6)
    c.value = 'SOA'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 7)
    c.value = 'NS'
    c.font = Font(bold=True)

    c = ws.cell(row=row, column=col + 8)
    c.value = 'Tags'
    c.font = Font(bold=True)

    row += 1

    for result in results:
        c = ws.cell(row=row, column=col)
        if result['domain']:
            c.value = result['domain']

        c = ws.cell(row=row, column=col + 1)
        if result['a_record']:
            if len(result['a_record']) > 1:
                c.value = ','.join(result['a_record'])
            else:
                c.value = result['a_record'][0]

        c = ws.cell(row=row, column=col + 2)
        if result['cname_record']:
            if len(result['cname_record']) > 1:
                c.value = ','.join(result['cname_record'])
            else:
                c.value = result['cname_record'][0]

        c = ws.cell(row=row, column=col + 3)
        if result['txt_record']:
            if len(result['txt_record']) > 1:
                c.value = ','.join(result['txt_record'])
            else:
                c.value = result['txt_record'][0]

        c = ws.cell(row=row, column=col + 4)
        if result['aaaa_record']:
            if len(result['aaaa_record']) > 1:
                c.value = ','.join(result['aaaa_record'])
            else:
                c.value = result['aaaa_record'][0]

        c = ws.cell(row=row, column=col + 5)
        if result['mx_record']:
            if len(result['mx_record']) > 1:
                c.value = ','.join(result['mx_record'])
            else:
                c.value = result['mx_record'][0]

        c = ws.cell(row=row, column=col + 6)
        if result['soa_record']:
            c.value = result['soa_record'][0]

        c = ws.cell(row=row, column=col + 7)
        if result['ns_record']:
            if len(result['ns_record']) > 1:
                c.value = ','.join(result['ns_record'])
            else:
                c.value = result['ns_record'][0]

        c = ws.cell(row=row, column=col + 8)
        if result['tags']:
            c.value = result['tags']
        row += 1
    fix_columns(ws)


def add_shodan_ws(orca_dbconn, wb):
    row = 1
    col = 1
    results = orca_dbconn.get_all_shodan_entries()

    ws1 = wb.active
    ws1.title = "SHODAN"

    c = ws1.cell(row=row, column=col)
    c.value = 'Asset ID'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 1)
    c.value = 'Host ID'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 2)
    c.value = 'IP Address'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 3)
    c.value = 'Hostname'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 4)
    c.value = 'Shodan_hostname'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 5)
    c.value = 'Network name'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 6)
    c.value = 'Country'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 7)
    c.value = 'ASN'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 8)
    c.value = 'Tags'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 9)
    c.value = 'Module'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 10)
    c.value = 'Port'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 11)
    c.value = 'CPE'
    c.font = Font(bold=True)

    c = ws1.cell(row=row, column=col + 12)
    c.value = 'SHODAN Banner'
    c.font = Font(bold=True)

    row += 1

    for result in results:
        c = ws1.cell(row=row, column=col)
        c.value = result['asset_id']

        c = ws1.cell(row=row, column=col + 1)
        c.value = result['host_id']

        c = ws1.cell(row=row, column=col + 2)
        c.value = result['ipaddr']

        c = ws1.cell(row=row, column=col + 3)

        if result['host_id'] == 0 or result['host_id'] == 999:
            c.value = '{}'
        else:
            response = orca_dbconn.get_hostname_from_hostid(result['host_id'])
            if not response:
                c.value = '{}'
            else:
                c.value = response[0]

        c = ws1.cell(row=row, column=col + 4)
        response = orca_dbconn.get_shodan_hostname_from_hostid(result['host_id'])
        if response and response is not None and len(response) > 0:
            if response[0] is not None:
                c.value = ','.join(response[0])

        c = ws1.cell(row=row, column=col + 5)
        c.value = result['netname']

        c = ws1.cell(row=row, column=col + 6)
        c.value = result['country']

        c = ws1.cell(row=row, column=col + 7)
        c.value = result['asn']

        if result['tags']:
            c = ws1.cell(row=row, column=col + 8)
            c.value = ', '.join(result['tags'])

        offset = col + 9

        same_port = False
        for ndx in range(len(result['cli'])):
            color = ''
            if (ndx % 2 == 0):
                color = "C0C0C0"
            else:
                color = "DDDDDD"
            c = ws1.cell(row=row, column=offset)
            # print("mods {}".format(result['cli']))
            c.value = result['cli'][ndx]
            c.fill = PatternFill("solid", fgColor=color)
            c.font = Font(bold=True)
            c = ws1.cell(row=row, column=offset + 1)

            if len(result['ports']) != len(result['cli']):
                same_port = True

            if same_port:
                c.value = (result['ports'][0])
            else:
                c.value = (result['ports'][ndx])

            c.fill = PatternFill("solid", fgColor=color)

            c = ws1.cell(row=row, column=offset + 2)
            c.fill = PatternFill("solid", fgColor=color)
            if 'cpe' in result['cpe']:
                # print("cpes: {}".format(result['cpe']))
                module_key = result['cli'][ndx]
                cpe = result['cpe']['cpe']
                for res in cpe:
                    if module_key in res:
                        c.value = str(res[module_key][0])

            c = ws1.cell(row=row, column=offset + 3)
            c.fill = PatternFill("solid", fgColor=color)
            if result['banner_shodan'][ndx]:
                banner = result['banner_shodan'][ndx]
                XLS_ILLEGAL_CHARS = r'[\000-\010]|[\013-\014]|[\016-\037]'
                banner = re.sub(XLS_ILLEGAL_CHARS, '?', banner)
                c.value = banner
                c.alignment = Alignment(wrap_text=False, shrink_to_fit=False)
            offset += 4
            # print('\n')
        row += 1
    fix_columns(ws1)


def create_xlsx(orca_dbconn, title, filename):
    wb = Workbook()
    add_shodan_ws(orca_dbconn, wb)
    add_dns_ws(orca_dbconn, wb)
    add_vuln_ws(orca_dbconn, wb)
    wb.save(filename)
    print("Wrote output to {}".format(filename))
